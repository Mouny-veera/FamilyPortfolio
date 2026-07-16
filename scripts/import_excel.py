#!/usr/bin/env python3
"""
Import existing Excel holdings into the portfolio database.

Usage:
    python3 scripts/import_excel.py <excel_file> <member_name>

Example:
    python3 scripts/import_excel.py ~/Downloads/"MOUNY Equity.xlsx" Mouny
"""

import sys
import re
import sqlite3
from datetime import date, datetime
from pathlib import Path

import openpyxl


DB_PATH = Path(__file__).resolve().parent.parent / "data" / "portfolio.db"

EQUITY_SHEET_PREFIX = "Equity"
PNL_SHEET_PREFIX = "P&L"

TICKER_NORMALIZE = {
    "INDUSIND BANK": "INDUSINDBK",
    "INOX WIND": "INOXWIND",
    "HINDUSTAN UNILEVER": "HINDUNILVR",
    "DEEPAK NITRITE": "DEEPAKNTR",
    "DEEPAK NITRATE": "DEEPAKNTR",
    "TATA CHEMICALS": "TATACHEM",
    "UNITED SPIRITS": "UNITDSPR",
    "TATA MOTORS  PV": "TATAMTRDVR",
    "TATA MOTORS  CV": "TATAMTRDVR",
    "HCL TECH": "HCLTECH",
    "KPIT TECH": "KPITTECH",
    "HINDUSTAN COPPER": "HINDCOPPER",
    "HINDUSTAN ZINC": "HINDZINC",
    "RELIANCE INDUSTRIES": "RELIANCE",
    "POWER GRID": "POWERGRID",
    "BAJAJ CONSUMER": "BAJAJCON",
    "ADANI POWER": "ADANIPOWER",
    "NATCO PHARMA": "NATCOPHARM",
    "VARUN BEVERAGES": "VBL",
    "BAJFINANCE": "BAJFINANCE",
    "BAJAJFINANCE": "BAJFINANCE",
    "ZYDUS LIFESCIENCES": "ZYDUSLIFE",
    "CAMLIN FINE SCIENCES": "CAMLINFINE",
    "JYOTHYLAB": "JYOTHYLAB",
    "JYOTHY LAB": "JYOTHYLAB",
    "IDFCFIRSTB": "IDFCFIRSTB",
    "SOUTHBANK": "SOUTHBANK",
    "TATA GOLD": "TATAGOLD",
    "DR REDDY": "DRREDDY",
    "EMAMI LTD": "EMAMILTD",
    "INOX WIND ": "INOXWIND",
    "PI INDUSTRIES": "PIIND",
    "APOLLO TYRES": "APOLLOTYRE",
    "BHARAT RASAYAN": "BHARATRAS",
    "VINATI ORGANICS": "VINATIORGA",
    "TATA TECHNOLOGIES": "TATATECH",
    "ARE&M": "ARE&M",
    "HINDALCO ": "HINDALCO",
    "GOODYEAR": "GOODYEAR",
}

SKIP_KEYWORDS = ["BONUS", "SPLIT", "NOW-", "NOW:-", "NOW:"]


def derive_financial_year(d: date) -> str:
    if d.month >= 4:
        return f"{d.year}-{str(d.year + 1)[2:]}"
    return f"{d.year - 1}-{str(d.year)[2:]}"


def parse_date(val) -> date | None:
    if isinstance(val, datetime):
        if val.year > 2100:
            return None
        return val.date()
    if isinstance(val, date):
        return val
    if not isinstance(val, str):
        return None

    val = val.strip()
    if not val:
        return None

    val = val.replace("//", "/")

    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(val, fmt).date()
            if parsed.year < 2020 or parsed.year > 2030:
                if fmt == "%d/%m/%y":
                    parts = val.split("/")
                    if len(parts) == 3:
                        yr = int(parts[2])
                        if yr < 100:
                            yr += 2000
                        try:
                            parsed = date(yr, int(parts[1]), int(parts[0]))
                        except ValueError:
                            continue
                        if 2020 <= parsed.year <= 2030:
                            return parsed
                continue
            return parsed
        except ValueError:
            continue

    return None


def normalize_ticker(raw: str) -> str:
    cleaned = raw.strip()
    if cleaned in TICKER_NORMALIZE:
        return TICKER_NORMALIZE[cleaned]
    cleaned = re.sub(r"\s+", "", cleaned).upper()
    return cleaned


def is_annotation_row(row_data: dict) -> bool:
    desc = str(row_data.get("D", ""))
    rate = row_data.get("F", "")

    if isinstance(rate, str) and any(kw in rate for kw in SKIP_KEYWORDS):
        return True

    for kw in ["BONUS", "SPLIT QTY", "split-", "split "]:
        if kw.lower() in desc.lower():
            return True

    return False


def get_member_id(conn: sqlite3.Connection, member_name: str) -> int:
    row = conn.execute(
        "SELECT id FROM members WHERE LOWER(name) = LOWER(?)", (member_name,)
    ).fetchone()
    if not row:
        names = [r[0] for r in conn.execute("SELECT name FROM members").fetchall()]
        print(f"ERROR: Member '{member_name}' not found. Available: {names}")
        sys.exit(1)
    return row[0]


def import_equity_sheet(conn: sqlite3.Connection, ws, member_id: int, fy: str):
    lots = []
    prev_ticker = None
    ticker_lot_count: dict[str, int] = {}

    for row in ws.iter_rows(min_row=2, max_col=13):
        vals = {c.column_letter: c.value for c in row if c.value is not None}
        if not vals:
            continue

        desc = str(vals.get("D", "")).strip()
        if not desc or "Description" in desc or "DATE" in desc:
            continue
        if vals.get("F") == "INVESTED" or vals.get("G") == "Buy Value":
            break

        if is_annotation_row(vals):
            continue

        qty = vals.get("E")
        rate = vals.get("F")
        if not isinstance(qty, (int, float)) or not isinstance(rate, (int, float)):
            continue

        buy_date = parse_date(vals.get("C"))
        if not buy_date:
            continue

        ticker = normalize_ticker(desc)
        buy_value = round(qty * rate, 2)

        if ticker not in ticker_lot_count:
            ticker_lot_count[ticker] = 0
        ticker_lot_count[ticker] += 1

        count = ticker_lot_count[ticker]
        if count == 1:
            lot_label = str(len([t for t in ticker_lot_count if ticker_lot_count[t] >= 1]))
        else:
            base = str(
                list(ticker_lot_count.keys()).index(ticker) + 1
            )
            suffix = chr(ord("A") + count - 2)
            lot_label = f"{base}{suffix}"

        lot_fy = derive_financial_year(buy_date)

        lots.append((
            member_id, ticker, buy_date.isoformat(), qty, rate, buy_value,
            lot_label, lot_fy, None,
        ))

    if lots:
        conn.executemany(
            """INSERT INTO lots
               (member_id, ticker, buy_date, buy_qty, buy_rate, buy_value,
                lot_label, financial_year, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            lots,
        )
        conn.commit()

    return len(lots)


def import_pnl_sheet(conn: sqlite3.Connection, ws, member_id: int, fy: str):
    records = []

    for row in ws.iter_rows(min_row=2, max_col=13):
        vals = {c.column_letter: c.value for c in row if c.value is not None}
        if not vals:
            continue

        desc = str(vals.get("D", "")).strip()
        if not desc or "Description" in desc:
            continue
        if vals.get("F") == "INVESTED" or vals.get("G") == "Buy Value":
            break

        if is_annotation_row(vals):
            continue

        buy_qty = vals.get("E")
        buy_rate = vals.get("F")
        sell_qty = vals.get("I")
        sell_rate = vals.get("J")

        if not all(isinstance(v, (int, float)) for v in [buy_qty, buy_rate, sell_qty, sell_rate]):
            continue

        buy_date = parse_date(vals.get("C"))
        sell_date = parse_date(vals.get("H"))
        if not buy_date:
            continue
        if not sell_date:
            sell_date = buy_date

        ticker = normalize_ticker(desc)
        buy_value = round(buy_qty * buy_rate, 2)
        sell_value = round(sell_qty * sell_rate, 2)
        profit_loss = round(sell_value - buy_value, 2)
        profit_loss_pct = round((profit_loss / buy_value) * 100, 2) if buy_value else 0.0

        sell_fy = derive_financial_year(sell_date)
        lot_label = str(vals.get("B", ""))

        records.append((
            member_id, ticker, buy_date.isoformat(), buy_qty, buy_rate, buy_value,
            sell_date.isoformat(), sell_qty, sell_rate, sell_value,
            profit_loss, profit_loss_pct, sell_fy, lot_label, None,
        ))

    if records:
        conn.executemany(
            """INSERT INTO realized_pnl
               (member_id, ticker, buy_date, buy_qty, buy_rate, buy_value,
                sell_date, sell_qty, sell_rate, sell_value,
                profit_loss, profit_loss_pct, financial_year, lot_label, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            records,
        )
        conn.commit()

    return len(records)


def extract_fy_from_sheet_name(name: str) -> str:
    match = re.search(r"(\d{4})-(\d{2,4})", name)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return ""


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    excel_path = Path(sys.argv[1]).expanduser()
    member_name = sys.argv[2]

    if not excel_path.exists():
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    if not DB_PATH.exists():
        print(f"ERROR: Database not found at {DB_PATH}")
        print("Start the backend first: cd backend && python -m uvicorn app.main:app")
        sys.exit(1)

    conn = sqlite3.connect(str(DB_PATH))
    member_id = get_member_id(conn, member_name)
    print(f"Importing for member: {member_name} (id={member_id})")

    existing_lots = conn.execute(
        "SELECT COUNT(*) FROM lots WHERE member_id = ?", (member_id,)
    ).fetchone()[0]
    existing_pnl = conn.execute(
        "SELECT COUNT(*) FROM realized_pnl WHERE member_id = ?", (member_id,)
    ).fetchone()[0]

    if existing_lots > 0 or existing_pnl > 0:
        print(f"WARNING: Member already has {existing_lots} lots and {existing_pnl} P&L records.")
        resp = input("Delete existing data and re-import? (y/N): ").strip().lower()
        if resp != "y":
            print("Aborted.")
            sys.exit(0)
        conn.execute("DELETE FROM lots WHERE member_id = ?", (member_id,))
        conn.execute("DELETE FROM realized_pnl WHERE member_id = ?", (member_id,))
        conn.commit()
        print("Cleared existing data.")

    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)
    wb = openpyxl.load_workbook(str(excel_path))

    equity_sheets = []
    pnl_sheets = []

    for name in wb.sheetnames:
        clean = name.strip()
        if clean.lower().startswith("p&l"):
            pnl_sheets.append(name)
        elif clean.lower().startswith("equity") and "OLD" not in name.upper():
            equity_sheets.append(name)

    most_recent_equity = equity_sheets[-1] if equity_sheets else None

    total_lots = 0
    total_pnl = 0

    if most_recent_equity:
        ws = wb[most_recent_equity]
        fy = extract_fy_from_sheet_name(most_recent_equity)
        print(f"\nImporting active holdings from: '{most_recent_equity}' (FY: {fy})")
        count = import_equity_sheet(conn, ws, member_id, fy)
        total_lots += count
        print(f"  -> {count} lots imported")

    for sheet_name in pnl_sheets:
        ws = wb[sheet_name]
        fy = extract_fy_from_sheet_name(sheet_name)
        print(f"\nImporting P&L from: '{sheet_name}' (FY: {fy})")
        count = import_pnl_sheet(conn, ws, member_id, fy)
        total_pnl += count
        print(f"  -> {count} P&L records imported")

    conn.close()
    print(f"\nDone! Imported {total_lots} active lots + {total_pnl} P&L records for {member_name}.")


if __name__ == "__main__":
    main()
