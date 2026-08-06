"""Format-agnostic Excel parsing for the holdings import.

Broker exports and hand-maintained workbooks put the same information in
different columns under different names, so nothing here is positional. We find
the header row, match each column to a field by its header text, fall back to
inferring from cell types, and attach a confidence to every guess so the UI can
ask the user to confirm anything shaky.

Pure parsing — no DB access, no writes. The router owns persistence.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field as dataclass_field
from datetime import date, datetime
from io import BytesIO
from typing import Any

import openpyxl

# Rows to scan when looking for the header — broker exports often carry a few
# title/account/disclaimer rows above it.
HEADER_SEARCH_ROWS = 25
# Cells sampled per column for type inference.
TYPE_SAMPLE_ROWS = 40

FIELDS = (
    "buy_date", "description", "qty", "buy_rate", "buy_value",
    "sell_date", "sell_qty", "sell_rate", "sell_value",
)

# Exact header wording seen across the family workbooks and common Indian broker
# exports. Matched after normalisation, so case/punctuation/spacing don't matter.
HEADER_SYNONYMS: dict[str, tuple[str, ...]] = {
    "buy_date": (
        "date", "buy date", "purchase date", "trade date", "transaction date",
        "date of purchase", "buy dt", "acquisition date", "entry date",
    ),
    "sell_date": (
        "sell date", "sale date", "exit date", "date of sale", "sell dt",
        "redemption date", "transfer date",
    ),
    "description": (
        "description", "scrip", "scrip name", "symbol", "stock", "stock name",
        "company", "company name", "security", "security name", "instrument",
        "particulars", "name", "isin description", "share name",
    ),
    "qty": (
        "qty", "quantity", "shares", "units", "no of shares", "buy qty",
        "buy quantity", "purchase qty", "purchase quantity",
        "quantity bought", "qty bought",
    ),
    "buy_rate": (
        "rate", "price", "buy rate", "buy price", "purchase price", "avg price",
        "average price", "cost", "cost price", "rate per share", "buy avg",
        "purchase rate", "acquisition cost",
    ),
    "sell_qty": (
        "sell qty", "sell quantity", "sale qty", "sale quantity",
        "quantity sold", "qty sold", "units sold",
    ),
    "sell_rate": (
        "sell rate", "sell price", "sale price", "sale rate", "selling price",
        "sell avg", "realisation price", "realization price",
    ),
    # Money columns are authoritative — see parse_sheet. A bonus or split makes
    # sell qty diverge from buy qty, so rates and quantities cannot be trusted
    # to reconstruct the trade, but the rupee values always reconcile.
    "buy_value": ("buy value", "purchase value", "invested", "cost value",
                  "buy amount", "purchase amount", "investment"),
    "sell_value": ("sell value", "sale value", "sell amount", "sale amount",
                   "realised value", "realized value", "proceeds"),
}

# The totals line that closes a sheet.
TERMINATOR_TOKENS = ("invested", "total", "grand total")

# Sheet-name hints for whether a tab holds open positions or closed trades.
REALIZED_SHEET_HINTS = ("p&l", "pnl", "realized", "realised", "capital gain",
                        "sold", "sale", "profit")
HOLDINGS_SHEET_HINTS = ("equity", "holding", "portfolio", "position", "current")

# Rows that annotate a corporate action rather than record a trade.
ANNOTATION_KEYWORDS = ("bonus", "split", "now-", "now:-", "now:", "demerger",
                       "consolidation")

# Company-name spellings from the family workbooks that don't reduce to their
# NSE symbol mechanically. Consulted only after an exact symbol lookup fails.
TICKER_ALIASES: dict[str, str] = {
    "INDUSIND BANK": "INDUSINDBK", "INOX WIND": "INOXWIND",
    "HINDUSTAN UNILEVER": "HINDUNILVR", "DEEPAK NITRITE": "DEEPAKNTR",
    "DEEPAK NITRATE": "DEEPAKNTR", "TATA CHEMICALS": "TATACHEM",
    "UNITED SPIRITS": "UNITDSPR",
    # Tata Motors demerged: passenger vehicles trade as TMPV, commercial as
    # TMCV. Both previously pointed at TATAMTRDVR, delisted with the 2024 DVR
    # conversion.
    "TATA MOTORS PV": "TMPV", "TATA MOTORS CV": "TMCV",
    "HCL TECH": "HCLTECH",
    "KPIT TECH": "KPITTECH", "HINDUSTAN COPPER": "HINDCOPPER",
    "HINDUSTAN ZINC": "HINDZINC", "RELIANCE INDUSTRIES": "RELIANCE",
    "POWER GRID": "POWERGRID", "BAJAJ CONSUMER": "BAJAJCON",
    "ADANI POWER": "ADANIPOWER", "NATCO PHARMA": "NATCOPHARM",
    "VARUN BEVERAGES": "VBL", "BAJAJFINANCE": "BAJFINANCE",
    "ZYDUS LIFESCIENCES": "ZYDUSLIFE", "CAMLIN FINE SCIENCES": "CAMLINFINE",
    "JYOTHY LAB": "JYOTHYLAB", "TATA GOLD": "TATAGOLD", "DR REDDY": "DRREDDY",
    "EMAMI LTD": "EMAMILTD", "PI INDUSTRIES": "PIIND",
    "APOLLO TYRES": "APOLLOTYRE", "BHARAT RASAYAN": "BHARATRAS",
    "VINATI ORGANICS": "VINATIORGA", "TATA TECHNOLOGIES": "TATATECH",
}

# Suffixes brokers append that aren't part of the NSE symbol.
COMPANY_SUFFIXES = re.compile(
    r"\b(limited|ltd|private|pvt|company|co|corporation|corp|inc|"
    r"industries|india|equity|shares?)\b\.?", re.I
)


@dataclass
class ColumnGuess:
    field: str
    index: int
    header: str | None
    confidence: float
    reason: str


@dataclass
class TickerMatch:
    status: str                     # exact | alias | fuzzy | unmatched
    ticker: str | None
    confidence: float
    candidates: list[dict[str, Any]] = dataclass_field(default_factory=list)


@dataclass
class DateRepair:
    field: str
    original: str
    repaired: date
    note: str


@dataclass
class ParsedRow:
    row_number: int
    raw_description: str
    match: TickerMatch
    buy_date: date | None = None
    qty: float | None = None
    buy_rate: float | None = None
    buy_value: float | None = None
    sell_date: date | None = None
    sell_qty: float | None = None
    sell_rate: float | None = None
    sell_value: float | None = None
    profit_loss: float | None = None
    repairs: list[DateRepair] = dataclass_field(default_factory=list)
    corporate_action: str | None = None


@dataclass
class SkippedRow:
    row_number: int
    reason: str
    preview: str


@dataclass
class Annotation:
    """A bonus/split note. Not a trade, but it explains a quantity change."""
    row_number: int
    text: str


@dataclass
class SheetPreview:
    name: str
    kind: str                       # holdings | realized
    header_row: int | None
    mapping: dict[str, ColumnGuess]
    rows: list[ParsedRow]
    skipped: list[SkippedRow]
    annotations: list[Annotation] = dataclass_field(default_factory=list)
    financial_year: str | None = None
    selected: bool = True
    skip_reason: str | None = None


def _norm_header(value: Any) -> str:
    """Lowercase, strip punctuation, collapse whitespace."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[._/\\()\[\]#*:]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _is_date(value: Any) -> bool:
    return isinstance(value, (datetime, date)) or parse_date(value) is not None


def parse_date(value: Any) -> date | None:
    """Parse the date spellings seen across brokers and hand-kept sheets."""
    if isinstance(value, datetime):
        return value.date() if 1990 < value.year < 2100 else None
    if isinstance(value, date):
        return value if 1990 < value.year < 2100 else None
    if not isinstance(value, str):
        return None

    text = value.strip().replace("//", "/").replace(".", "/").replace("-", "/")
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d", "%m/%d/%Y", "%d/%b/%Y", "%d/%B/%Y"):
        try:
            parsed = datetime.strptime(text, fmt).date()
        except ValueError:
            continue
        if 1990 < parsed.year < 2100:
            return parsed
    return None


def repair_date(value: Any, *, plausible_from: int = 2000,
                plausible_to: int = 2100) -> tuple[date, str] | None:
    """Recover a date from a typo'd string.

    Real workbooks accumulate slips — a dropped separator, a mistyped year. The
    rows are otherwise valid trades, so rejecting them loses real holdings.
    Every repair is reported so the user confirms it rather than trusting us.
    """
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None

    parts = [p for p in re.split(r"[/\-.\s]+", text) if p]
    candidates: list[tuple[str, str]] = []

    # Separator dropped between month and year: "03/112025" -> 03/11/2025
    if len(parts) == 2 and len(parts[1]) == 6 and parts[1].isdigit():
        candidates.append((f"{parts[0]}/{parts[1][:2]}/{parts[1][2:]}",
                           "inserted missing separator"))
    # Whole date run together: "03112025"
    if len(parts) == 1 and len(parts[0]) == 8 and parts[0].isdigit():
        blob = parts[0]
        candidates.append((f"{blob[:2]}/{blob[2:4]}/{blob[4:]}",
                           "split run-together digits"))
    # Year missing a digit: "206"/"0205" -> 2026/2025. Digits are in order with
    # one dropped, so reinstate it before the final digit.
    if len(parts) == 3 and parts[2].isdigit():
        year = parts[2].lstrip("0")
        if len(year) == 3 and year.startswith("20"):
            candidates.append((f"{parts[0]}/{parts[1]}/20{year[1]}{year[2]}"
                               .replace("200", "202", 1),
                               f"year {parts[2]!r} read as 202{year[2]}"))
            candidates.append((f"{parts[0]}/{parts[1]}/202{year[2]}",
                               f"year {parts[2]!r} read as 202{year[2]}"))

    for text_candidate, note in candidates:
        parsed = parse_date(text_candidate)
        if parsed and plausible_from <= parsed.year <= plausible_to:
            return parsed, note
    return None


def _is_number(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return value == value          # reject NaN
    if isinstance(value, str):
        try:
            float(value.replace(",", "").strip())
            return True
        except ValueError:
            return False
    return False


def _to_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if value == value else None
    if isinstance(value, str):
        try:
            return float(value.replace(",", "").strip())
        except ValueError:
            return None
    return None


def detect_header_row(grid: list[list[Any]]) -> int | None:
    """Index of the row whose cells best match known field headers."""
    best_idx, best_score = None, 0.0
    for idx, row in enumerate(grid[:HEADER_SEARCH_ROWS]):
        hits = 0
        for cell in row:
            norm = _norm_header(cell)
            if not norm:
                continue
            for synonyms in HEADER_SYNONYMS.values():
                if norm in synonyms:
                    hits += 1
                    break
        # Require at least a date-ish and a name-ish column to call it a header.
        if hits >= 2 and hits > best_score:
            best_idx, best_score = idx, hits
    return best_idx


def _match_header_to_field(norm: str) -> tuple[str, float] | None:
    """Exact synonym match first, then token containment at lower confidence."""
    for fld, synonyms in HEADER_SYNONYMS.items():
        if norm in synonyms:
            return fld, 1.0

    # A header that names its side ("Sell Quantity") must not be captured by a
    # generic token ("quantity") belonging to the other side.
    sells = norm.startswith(("sell", "sale")) or " sold" in norm
    buys = norm.startswith(("buy", "purchase")) or " bought" in norm

    best: tuple[str, float] | None = None
    best_len = 0
    for fld, synonyms in HEADER_SYNONYMS.items():
        is_sell_field = fld.startswith("sell")
        if sells and not is_sell_field and fld != "description":
            continue
        if buys and is_sell_field:
            continue
        for syn in synonyms:
            if len(syn) > 3 and (syn in norm or norm in syn) and len(syn) > best_len:
                best, best_len = (fld, 0.65), len(syn)
    return best


def _infer_by_type(column: list[Any]) -> tuple[str, float] | None:
    """Classify a column from its cell types when the header tells us nothing."""
    values = [v for v in column if v is not None and str(v).strip() != ""]
    if len(values) < 3:
        return None
    n = len(values)
    dates = sum(1 for v in values if _is_date(v))
    numbers = [v for v in values if _is_number(v)]
    strings = sum(1 for v in values if isinstance(v, str) and not _is_number(v)
                  and not _is_date(v))

    if dates / n > 0.7:
        return "buy_date", 0.5
    if strings / n > 0.7:
        return "description", 0.5
    if len(numbers) / n > 0.7:
        nums = [_to_number(v) for v in numbers]
        nums = [x for x in nums if x is not None]
        if not nums:
            return None
        whole = sum(1 for x in nums if abs(x - round(x)) < 1e-9) / len(nums)
        # Quantities are whole numbers; rates carry paise.
        return ("qty", 0.45) if whole > 0.85 else ("buy_rate", 0.45)
    return None


def detect_columns(grid: list[list[Any]], header_row: int | None) -> dict[str, ColumnGuess]:
    """Map fields to column indexes via headers, then type inference."""
    width = max((len(r) for r in grid), default=0)
    mapping: dict[str, ColumnGuess] = {}

    if header_row is not None:
        header_cells = grid[header_row]
        for idx in range(min(width, len(header_cells))):
            norm = _norm_header(header_cells[idx])
            if not norm:
                continue
            hit = _match_header_to_field(norm)
            if not hit:
                continue
            fld, conf = hit
            existing = mapping.get(fld)
            if existing is None or conf > existing.confidence:
                mapping[fld] = ColumnGuess(
                    field=fld, index=idx, header=str(header_cells[idx]).strip(),
                    confidence=conf, reason=f"header {header_cells[idx]!r}",
                )

    body_start = (header_row + 1) if header_row is not None else 0
    body = grid[body_start:body_start + TYPE_SAMPLE_ROWS]
    taken = {g.index for g in mapping.values()}

    for idx in range(width):
        if idx in taken:
            continue
        column = [row[idx] if idx < len(row) else None for row in body]
        hit = _infer_by_type(column)
        if not hit:
            continue
        fld, conf = hit
        if fld in mapping:
            continue
        mapping[fld] = ColumnGuess(
            field=fld, index=idx, header=None, confidence=conf,
            reason="inferred from cell types",
        )
        taken.add(idx)

    return mapping


def classify_sheet(name: str, mapping: dict[str, ColumnGuess],
                   body: list[list[Any]] | None = None) -> str:
    """Open positions or closed trades.

    The sheet name decides it. Both sheet types declare the same sell columns —
    a holdings sheet simply leaves them empty — so the presence of a sell column
    proves nothing; only populated sell cells do.
    """
    lowered = name.strip().lower()
    if any(h in lowered for h in REALIZED_SHEET_HINTS):
        return "realized"
    if any(h in lowered for h in HOLDINGS_SHEET_HINTS):
        return "holdings"

    guess = mapping.get("sell_date")
    if guess is not None and body:
        filled = sum(
            1 for row in body
            if guess.index < len(row) and row[guess.index] is not None
            and str(row[guess.index]).strip() != ""
        )
        if filled:
            return "realized"
    return "holdings"


def extract_financial_year(name: str) -> str | None:
    match = re.search(r"(\d{4})\s*-\s*(\d{2,4})", name)
    if not match:
        return None
    start, end = match.group(1), match.group(2)
    return f"{start}-{end[-2:]}"


def normalize_company_name(raw: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(raw)).strip()
    cleaned = COMPANY_SUFFIXES.sub("", cleaned)
    return re.sub(r"\s+", " ", cleaned).strip().upper()


def resolve_ticker(raw: str, *, symbol_set: set[str] | None = None,
                   fuzzy=None) -> TickerMatch:
    """Description -> NSE symbol.

    Exact symbol match, then the alias table, then fuzzy candidates for the user
    to confirm. Never guesses silently: anything below an exact/alias hit comes
    back for confirmation.
    """
    if symbol_set is None:
        from .nse_master import get_nse_symbol_set
        symbol_set = get_nse_symbol_set()
    if fuzzy is None:
        from .nse_master import fuzzy_match_ticker
        fuzzy = fuzzy_match_ticker

    cleaned = normalize_company_name(raw)
    compact = re.sub(r"[^A-Z0-9&]", "", cleaned)

    if compact and compact in symbol_set:
        return TickerMatch("exact", compact, 1.0)

    # An alias is only as good as the symbol it points at. Listings change —
    # companies demerge, DVRs convert — so a mapping written once can rot. Check
    # the target still exists rather than trusting it; a stale alias then
    # surfaces for confirmation instead of failing at commit.
    for key in (cleaned, str(raw).strip().upper()):
        alias = TICKER_ALIASES.get(key)
        if alias and alias in symbol_set:
            return TickerMatch("alias", alias, 0.95)

    try:
        candidates = fuzzy(compact or cleaned, top_n=5) or []
    except Exception:
        candidates = []

    if candidates:
        return TickerMatch("fuzzy", None, 0.0, candidates)
    return TickerMatch("unmatched", None, 0.0, [])


def _is_annotation(cells: list[Any]) -> bool:
    joined = " ".join(str(c) for c in cells if c is not None).lower()
    return any(kw in joined for kw in ANNOTATION_KEYWORDS)


def parse_sheet(name: str, grid: list[list[Any]], *, resolver=resolve_ticker) -> SheetPreview:
    header_row = detect_header_row(grid)
    mapping = detect_columns(grid, header_row)
    start = (header_row + 1) if header_row is not None else 0
    kind = classify_sheet(name, mapping, grid[start:])

    rows: list[ParsedRow] = []
    skipped: list[SkippedRow] = []
    annotations: list[Annotation] = []

    def cell(row: list[Any], fld: str) -> Any:
        guess = mapping.get(fld)
        if guess is None or guess.index >= len(row):
            return None
        return row[guess.index]

    def read_date(row: list[Any], fld: str, repairs: list[DateRepair]):
        raw = cell(row, fld)
        parsed = parse_date(raw)
        if parsed is not None:
            return parsed
        fixed = repair_date(raw)
        if fixed is None:
            return None
        repaired, note = fixed
        repairs.append(DateRepair(fld, str(raw).strip(), repaired, note))
        return repaired

    for offset, row in enumerate(grid[start:]):
        row_number = start + offset + 1        # 1-based, matches Excel
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue

        preview = " | ".join(str(c) for c in row[:6] if c is not None)[:120]

        # Totals line closes the sheet; everything after it is summary.
        joined = " ".join(str(c) for c in row if c is not None).strip().lower()
        if any(joined.startswith(t) or f" {t}" in joined[:40] for t in TERMINATOR_TOKENS):
            if cell(row, "description") in (None, ""):
                break

        # Bonus/split notes aren't trades, but they explain why a quantity
        # changed — keep them visible instead of dropping them silently.
        if _is_annotation(row):
            text = " ".join(str(c) for c in row if c is not None).strip()
            annotations.append(Annotation(row_number, text[:160]))
            continue

        desc = cell(row, "description")
        if desc is None or not str(desc).strip():
            skipped.append(SkippedRow(row_number, "no description", preview))
            continue

        qty = _to_number(cell(row, "qty"))
        rate = _to_number(cell(row, "buy_rate"))
        if qty is None or rate is None:
            skipped.append(SkippedRow(row_number, "missing quantity or rate", preview))
            continue

        repairs: list[DateRepair] = []
        buy_date = read_date(row, "buy_date", repairs)
        if buy_date is None:
            skipped.append(SkippedRow(row_number, "unreadable buy date", preview))
            continue

        # Money is authoritative. Fall back to qty x rate only when the sheet
        # doesn't carry a value column.
        buy_value = _to_number(cell(row, "buy_value"))
        if buy_value is None:
            buy_value = round(qty * rate, 2)

        sell_date = read_date(row, "sell_date", repairs)
        sell_qty = _to_number(cell(row, "sell_qty"))
        sell_rate = _to_number(cell(row, "sell_rate"))
        sell_value = _to_number(cell(row, "sell_value"))
        if sell_value is None and sell_qty is not None and sell_rate is not None:
            sell_value = round(sell_qty * sell_rate, 2)

        # A realized row needs a sell date — it decides the P&L financial year.
        if kind == "realized" and sell_value is not None and sell_date is None:
            skipped.append(SkippedRow(row_number, "unreadable sell date", preview))
            continue

        profit_loss = (round(sell_value - buy_value, 2)
                       if sell_value is not None else None)

        action = None
        if sell_qty is not None and qty and sell_qty != qty:
            ratio = sell_qty / qty
            action = (f"quantity changed {qty:g} -> {sell_qty:g} "
                      f"({ratio:g}x) — likely bonus or split")

        rows.append(ParsedRow(
            row_number=row_number,
            raw_description=str(desc).strip(),
            match=resolver(str(desc)),
            buy_date=buy_date,
            qty=qty,
            buy_rate=rate,
            buy_value=buy_value,
            sell_date=sell_date,
            sell_qty=sell_qty,
            sell_rate=sell_rate,
            sell_value=sell_value,
            profit_loss=profit_loss,
            repairs=repairs,
            corporate_action=action,
        ))

    return SheetPreview(
        name=name, kind=kind, header_row=header_row, mapping=mapping,
        rows=rows, skipped=skipped, annotations=annotations,
        financial_year=extract_financial_year(name),
    )


def load_grid(ws, max_rows: int = 5000, max_cols: int = 40) -> list[list[Any]]:
    grid: list[list[Any]] = []
    for row in ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
        grid.append(list(row))
    return grid


def parse_workbook(data: bytes, *, resolver=resolve_ticker) -> list[SheetPreview]:
    """Parse every readable sheet. Caller enforces the upload size limit."""
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        previews = []
        for name in wb.sheetnames:
            grid = load_grid(wb[name])
            if not grid:
                continue
            preview = parse_sheet(name, grid, resolver=resolver)
            # Sheets with no usable rows and no detected columns are noise
            # (cover pages, disclaimers) — don't surface them.
            if preview.rows or preview.mapping:
                previews.append(preview)
        return previews
    finally:
        wb.close()


def select_import_sheets(previews: list[SheetPreview]) -> list[SheetPreview]:
    """Mark which sheets actually import.

    Holdings sheets are cumulative snapshots — each year restates every open lot,
    so importing more than the newest duplicates the portfolio. Realized sheets
    are the opposite: each records the trades closed in its own year and they
    never overlap, so all of them import.
    """
    holdings = [p for p in previews if p.kind == "holdings"]

    def sort_key(p: SheetPreview) -> tuple[int, str]:
        fy = p.financial_year
        return (1, fy) if fy else (0, p.name)

    latest = max(holdings, key=sort_key) if holdings else None

    for preview in previews:
        if preview.kind != "holdings":
            preview.selected = True
            continue
        if preview is latest:
            preview.selected = True
        else:
            preview.selected = False
            preview.skip_reason = (
                f"superseded by {latest.name.strip()!r} — holdings sheets restate "
                f"every open lot, so importing both would duplicate them"
            )
    return previews


def derive_financial_year(d: date) -> str:
    """Apr 1+ belongs to {year}-{year+1}; before Apr to {year-1}-{year}."""
    return f"{d.year}-{str(d.year + 1)[2:]}" if d.month >= 4 else f"{d.year - 1}-{str(d.year)[2:]}"


def assign_lot_labels(tickers: list[str]) -> list[str]:
    """Sub-lots per ticker: 1, 1A, 1B — the family's lot convention."""
    order: dict[str, int] = {}
    seen: dict[str, int] = {}
    labels = []
    for ticker in tickers:
        if ticker not in order:
            order[ticker] = len(order) + 1
            seen[ticker] = 0
        seen[ticker] += 1
        base = order[ticker]
        n = seen[ticker]
        labels.append(str(base) if n == 1 else f"{base}{chr(ord('A') + n - 2)}")
    return labels
